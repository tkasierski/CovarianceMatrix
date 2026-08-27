from pathlib import Path

import openpyxl
import pandas as pd

from covariance_matrix.analytics import analyze_returns
from covariance_matrix.excel import build_workbook


def _formula_cells(workbook: openpyxl.Workbook, sheet_name: str) -> list[str]:
    worksheet = workbook[sheet_name]
    return [
        cell.value
        for row in worksheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]


def test_workbook_uses_direct_ranges_and_dedicated_risk_free_rate(tmp_path: Path):
    returns = pd.DataFrame(
        {"A": [0.01, 0.02, -0.01], "B": [0.02, 0.01, 0.00]},
        index=pd.date_range("2024-01-31", periods=3, freq="ME"),
    )
    result = analyze_returns(returns, min_observations=2)
    output = tmp_path / "test.xlsx"
    build_workbook(result, output, risk_free_rate=0.04)
    workbook = openpyxl.load_workbook(output, data_only=False)

    covariance = workbook["Covariance_Matrix"]
    covariance_formula = covariance["B3"].value
    assert covariance["A1"].value == "Monthly Covariance Matrix"
    assert "INDEX" not in covariance_formula
    assert "MATCH" not in covariance_formula
    assert "Monthly_Simple_Returns" in covariance_formula
    assert "_xlfn.COVARIANCE.S" in covariance_formula

    dashboard = workbook["Portfolio_Dashboard"]
    assert dashboard["B3"].value == 0.04
    sharpe_formulas = [
        formula
        for formula in _formula_cells(workbook, "Portfolio_Dashboard")
        if "IFERROR((B" in formula
    ]
    assert sharpe_formulas and "$B$3" in sharpe_formulas[0]


def test_live_downside_and_drawdown_formulas_reference_returns(tmp_path: Path):
    returns = pd.DataFrame(
        {"A": [0.10, -0.20, 0.05, 0.25], "B": [0.02, -0.01, 0.03, -0.04]},
        index=pd.date_range("2024-01-31", periods=4, freq="ME"),
    )
    output = tmp_path / "test.xlsx"
    build_workbook(
        analyze_returns(returns, min_observations=2),
        output,
        minimum_acceptable_return=0.005,
    )
    workbook = openpyxl.load_workbook(output, data_only=False)
    downside = workbook["Downside_Risk_Metrics"]

    assert downside["B2"].value.startswith("=COUNT(")
    assert "STDEV.S" in downside["C2"].value
    assert "SUMPRODUCT" in downside["E2"].value
    assert "Portfolio_Dashboard'!$B$4" in downside["E2"].value
    assert "PERCENTILE" in downside["G2"].value
    assert "SUMIF" in downside["H2"].value
    assert "Drawdown_Series" in downside["K2"].value
    assert "MAX" in downside["L2"].value
    assert "INDEX" in downside["M2"].value

    drawdown_formulas = _formula_cells(workbook, "Drawdown_Series")
    assert drawdown_formulas
    assert any("Monthly_Simple_Returns" in formula for formula in drawdown_formulas)


def test_portfolio_backtest_is_live_and_date_windowed(tmp_path: Path):
    returns = pd.DataFrame(
        {"A": [0.10, -0.20, 0.05, 0.25], "B": [0.02, -0.01, 0.03, -0.04]},
        index=pd.date_range("2024-01-31", periods=4, freq="ME"),
    )
    output = tmp_path / "test.xlsx"
    build_workbook(analyze_returns(returns, min_observations=2), output)
    workbook = openpyxl.load_workbook(output, data_only=False)

    assert "Portfolio_Backtest" in workbook.sheetnames
    dashboard = workbook["Portfolio_Dashboard"]
    backtest = workbook["Portfolio_Backtest"]

    assert dashboard["D1"].value == "Backtest Start Date"
    assert dashboard["D2"].value == "Backtest End Date"
    assert dashboard["E1"].value == returns.index.min().to_pydatetime()
    assert dashboard["E2"].value == returns.index.max().to_pydatetime()

    assert "Portfolio_Dashboard'!$E$1" in backtest["B2"].value
    assert "Portfolio_Dashboard'!$E$2" in backtest["B2"].value
    assert "Monthly_Simple_Returns'!B2*'Portfolio_Dashboard'!$C$7" in backtest["C2"].value
    assert "Monthly_Simple_Returns'!C2*'Portfolio_Dashboard'!$C$8" in backtest["C2"].value
    assert "MAX" in backtest["E3"].value
    assert "/E" in backtest["F2"].value

    dashboard_formulas = _formula_cells(workbook, "Portfolio_Dashboard")
    backtest_metric_formulas = [
        formula for formula in dashboard_formulas if "Portfolio_Backtest" in formula
    ]
    assert any("PERCENTILE" in formula for formula in backtest_metric_formulas)
    assert any("SUMIF" in formula for formula in backtest_metric_formulas)
    assert any("MIN('Portfolio_Backtest'!$F$2" in formula for formula in backtest_metric_formulas)
    assert any("LOOKUP" in formula for formula in backtest_metric_formulas)


def test_active_formulas_do_not_contain_implicit_intersection_operator(tmp_path: Path):
    returns = pd.DataFrame(
        {"A": [0.01, 0.02, -0.01], "B": [0.02, 0.01, 0.00]},
        index=pd.date_range("2024-01-31", periods=3, freq="ME"),
    )
    output = tmp_path / "test.xlsx"
    build_workbook(analyze_returns(returns, min_observations=2), output)
    workbook = openpyxl.load_workbook(output, data_only=False)

    active_sheets = [
        "Covariance_Matrix",
        "Correlation_Matrix",
        "Portfolio_Dashboard",
        "Portfolio_Backtest",
        "Downside_Risk_Metrics",
        "Drawdown_Series",
    ]
    formulas = [
        formula
        for sheet_name in active_sheets
        for formula in _formula_cells(workbook, sheet_name)
    ]
    assert formulas
    assert all("@" not in formula for formula in formulas)


def test_validation_sheet_exists(tmp_path: Path):
    returns = pd.DataFrame(
        {"A": [0.01, 0.02]},
        index=pd.date_range("2024-01-31", periods=2, freq="ME"),
    )
    output = tmp_path / "test.xlsx"
    build_workbook(analyze_returns(returns, min_observations=2), output)
    workbook = openpyxl.load_workbook(output)
    assert "Validation" in workbook.sheetnames
    assert "Observation_Counts" in workbook.sheetnames
