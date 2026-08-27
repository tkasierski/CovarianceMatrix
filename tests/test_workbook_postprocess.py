from pathlib import Path

import openpyxl
import pandas as pd

from covariance_matrix.analytics import analyze_returns
from covariance_matrix.excel import build_workbook
from covariance_matrix.workbook_postprocess import apply_phase1_backtest_tweaks


def _find_label_row(worksheet, label: str) -> int:
    for row in range(1, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=1).value == label:
            return row
    raise AssertionError(f"Missing label: {label}")


def test_phase1_tweaks_fix_downside_and_add_annual_tail_risk(tmp_path: Path):
    returns = pd.DataFrame(
        {
            "A": [0.01, -0.02, 0.03, 0.01, -0.01, 0.02, 0.01, -0.03, 0.04, 0.01, -0.02, 0.02, 0.01],
            "B": [0.02, 0.01, -0.01, 0.00, 0.02, -0.02, 0.03, 0.01, -0.01, 0.02, 0.01, -0.02, 0.03],
        },
        index=pd.date_range("2023-01-31", periods=13, freq="ME"),
    )
    output = tmp_path / "test.xlsx"
    build_workbook(analyze_returns(returns, min_observations=2), output)
    apply_phase1_backtest_tweaks(output)

    workbook = openpyxl.load_workbook(output, data_only=False)
    dashboard = workbook["Portfolio_Dashboard"]
    backtest = workbook["Portfolio_Backtest"]

    downside_row = _find_label_row(dashboard, "Annualized Downside Deviation")
    downside_formula = dashboard.cell(row=downside_row, column=2).value
    assert "IFERROR((\'Portfolio_Backtest\'!$C$2" not in downside_formula
    assert "SUMPRODUCT(IFERROR" in downside_formula

    assert _find_label_row(dashboard, "Historical Monthly 95% VaR")
    assert _find_label_row(dashboard, "Historical Monthly 95% CVaR")
    assert _find_label_row(dashboard, "Historical Monthly 99% VaR")
    assert _find_label_row(dashboard, "Historical Monthly 99% CVaR")

    assert backtest["H1"].value == "Rolling 12M Portfolio Return"
    assert backtest["H13"].value.startswith("=IF(COUNT(C2:C13)=12,")
    assert "(1+C2)*(1+C3)" in backtest["H13"].value

    annual_var_row = _find_label_row(dashboard, "Historical Annual 95% VaR (Rolling 12M)")
    annual_cvar_row = _find_label_row(dashboard, "Historical Annual 95% CVaR (Rolling 12M)")
    assert "PERCENTILE('Portfolio_Backtest'!$H$2:$H$14,0.05)" in dashboard.cell(row=annual_var_row, column=2).value
    assert "SUMIF('Portfolio_Backtest'!$H$2:$H$14" in dashboard.cell(row=annual_cvar_row, column=2).value
