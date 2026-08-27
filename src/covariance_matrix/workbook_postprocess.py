from __future__ import annotations

from pathlib import Path

import openpyxl


def _find_label_row(worksheet, label: str) -> int:
    for row in range(1, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=1).value == label:
            return row
    raise ValueError(f"Could not find dashboard label: {label}")


def apply_phase1_backtest_tweaks(output_file: str | Path) -> None:
    """Apply Excel-only Phase I backtest fixes after the workbook is generated.

    The resulting workbook remains fully formula-driven. This post-processing step
    only writes additional formulas/labels into the generated XLSX file.
    """
    output_file = Path(output_file)
    workbook = openpyxl.load_workbook(output_file, data_only=False)
    dashboard = workbook["Portfolio_Dashboard"]
    backtest = workbook["Portfolio_Backtest"]

    first_data_row = 2
    last_data_row = backtest.max_row
    monthly_returns = f"'Portfolio_Backtest'!$C${first_data_row}:$C${last_data_row}"

    # The original SUMPRODUCT expression still evaluates arithmetic against the
    # formula-generated empty strings outside the selected backtest window.
    # Wrapping the per-observation downside expression in IFERROR converts those
    # non-numeric rows to zero while COUNT keeps the denominator numeric-only.
    downside_row = _find_label_row(dashboard, "Annualized Downside Deviation")
    dashboard.cell(row=downside_row, column=2).value = (
        f'=IFERROR(SQRT(SUMPRODUCT(IFERROR(({monthly_returns}<$B$4)*'
        f'({monthly_returns}-$B$4)^2,0))/COUNT({monthly_returns}))*SQRT(12),"")'
    )

    # Make the existing historical tail-risk horizon explicit.
    label_updates = {
        "Historical 95% VaR": "Historical Monthly 95% VaR",
        "Historical 95% CVaR": "Historical Monthly 95% CVaR",
        "Historical 99% VaR": "Historical Monthly 99% VaR",
        "Historical 99% CVaR": "Historical Monthly 99% CVaR",
    }
    for old_label, new_label in label_updates.items():
        row = _find_label_row(dashboard, old_label)
        dashboard.cell(row=row, column=1).value = new_label

    # Build overlapping rolling 12-month compounded portfolio returns inside the
    # selected backtest window. A value is produced only when all 12 monthly
    # portfolio returns are numeric, preventing partial-year observations.
    annual_return_col = 8  # H
    backtest.cell(row=1, column=annual_return_col).value = "Rolling 12M Portfolio Return"
    for row in range(first_data_row, last_data_row + 1):
        if row < first_data_row + 11:
            backtest.cell(row=row, column=annual_return_col).value = '=""'
            continue
        start_row = row - 11
        monthly_window = f"C{start_row}:C{row}"
        backtest.cell(row=row, column=annual_return_col).value = (
            f'=IF(COUNT({monthly_window})=12,PRODUCT(1+{monthly_window})-1,"")'
        )
        backtest.cell(row=row, column=annual_return_col).number_format = "0.00%"

    annual_returns = f"'Portfolio_Backtest'!$H${first_data_row}:$H${last_data_row}"
    best_month_row = _find_label_row(dashboard, "Best Month")
    start_metric_row = best_month_row + 2
    annual_metrics = [
        (
            "Historical Annual 95% VaR (Rolling 12M)",
            f'=IFERROR(-PERCENTILE({annual_returns},0.05),"")',
        ),
        (
            "Historical Annual 95% CVaR (Rolling 12M)",
            f'=IFERROR(-SUMIF({annual_returns},"<="&PERCENTILE({annual_returns},0.05),'
            f'{annual_returns})/COUNTIF({annual_returns},"<="&PERCENTILE({annual_returns},0.05)),"")',
        ),
        (
            "Historical Annual 99% VaR (Rolling 12M)",
            f'=IFERROR(-PERCENTILE({annual_returns},0.01),"")',
        ),
        (
            "Historical Annual 99% CVaR (Rolling 12M)",
            f'=IFERROR(-SUMIF({annual_returns},"<="&PERCENTILE({annual_returns},0.01),'
            f'{annual_returns})/COUNTIF({annual_returns},"<="&PERCENTILE({annual_returns},0.01)),"")',
        ),
    ]

    # Match the existing dashboard header/value formatting where possible.
    reference_label = dashboard.cell(row=best_month_row, column=1)
    reference_value = dashboard.cell(row=best_month_row, column=2)
    for offset, (label, formula) in enumerate(annual_metrics):
        row = start_metric_row + offset
        label_cell = dashboard.cell(row=row, column=1)
        value_cell = dashboard.cell(row=row, column=2)
        label_cell.value = label
        value_cell.value = formula
        if reference_label.has_style:
            label_cell._style = reference_label._style
        if reference_value.has_style:
            value_cell._style = reference_value._style
        value_cell.number_format = "0.00%"

    backtest.column_dimensions["H"].width = 24
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(output_file)


__all__ = ["apply_phase1_backtest_tweaks"]
